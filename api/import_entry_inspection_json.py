#
# Copyright (C) 2021 AssetOwl Technologies Pty Ltd, all rights reserved.
#

import json

import requests
import logging
import os
import openpyxl
from openpyxl.cell import cell

parent_dir = r'C:\Users\nuwan\Documents\test org'

wrkbk = openpyxl.load_workbook(r"C:\Users\nuwan\Downloads\test.xlsx")

sh = wrkbk.active

logging.basicConfig(filename='bulk_import.log' ,level=logging.DEBUG, format='%(asctime)s: %(levelname)s:%(message)s')

# base_url = 'https://app.pip.local'
#base_url = 'https://app.dev.internal.assetowl.com'
# base_url = 'https://app.test.internal.assetowl.com'
base_url = 'https://app.prod.assetowl.com'

session = requests.Session()
# session.verify = False  # when using local environment without valid SSL certificate


def login(username, password):
    login_json = {
        'username': username,
        'password': password,
        'deviceId': 'SCRIPT'
    }
    login_response = session.post(f'{base_url}/api/auth/login', data=json.dumps(login_json))
    login_response.raise_for_status()
    return login_response.json()


access_token = login('nuwan+support@assetowl.com', "!tYcE/'6jFUj@Dk@")['accessToken']

session.headers = {
    'Authorization': f'Bearer {access_token}'
}

org_id = 'c8f53258-7d9c-4f92-af49-28d36bfb541b'


#documents = db.fetch_all_unimported_documents(org_id)

for i in range(2, sh.max_row+1):
    
    address = sh.cell(row=i, column=4).value.replace("/","_")
    property_id = sh.cell(row=i, column=6).value
    path_to_json = rf'{parent_dir}\{address}\completed json'

    for root, dirs, files in os.walk(path_to_json):
        for filename in files:
            if filename.endswith(".json"):
                print(filename)
                path = os.path.join(root, filename)

                logging.info(f"*-*-*-*-*-*-*-*-*-*-*-* Started {address} import *-*-*-*-*-*-*-*-*-*-*-*")
                try:
                    with open(path, 'r', encoding='utf-8') as file:
                        template_json = file.read()
                except Exception as e:
                    logging.error(e)
                    continue              

                try:
                    session.post(f'{base_url}/api/import/orgs/{org_id}/properties/{property_id}/inspections', data=template_json, headers={'Content-Type': 'application/json'}).raise_for_status()
                    logging.info(f"Imported {address} successfully")
                except Exception as e:
                    logging.error(e)
                    continue
  

