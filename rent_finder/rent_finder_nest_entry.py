from re import split
import sys
sys.path.insert(0, r'C:\Users\Nuwan\Downloads\pdf_to_json-20210617T052048Z-001\pdf_to_json\report_conversion')
import Entry 
import openpyxl
from openpyxl.cell import cell
import os


directory = r"E:\Google Drive - AssetOwl\Test PCR data - rentfindinspector"

for root, subdirectories, files in os.walk(directory):
    for file in files:
        if file.endswith(".xlsx"):
            print(root)
            Entry.Entry.load_excel_files(root, 'c8f53258-7d9c-4f92-af49-28d36bfb541b')

            #start and end of the report
            sor = 'PLEASE NOTE Any amendments to this report must be listed in writing and a signed copy returned to the Managing Agents within SEVEN (7) days of receiving same. Failure to do this will result in the bond inspection being carried out against this original report.'
            eor = 'Approximate dates when work was last done on Residential Premises:'


            for file in Entry.Entry.files:
                wb = openpyxl.load_workbook(rf'{file}')
                sh = wb.active

                entry = Entry.Entry('ENTRY')
                is_start = False

                # iterate through excel and extract data
                for i in range(3, sh.max_row+1):
                    title = sh.cell(row=i, column=1)
                    clean = sh.cell(row=i, column=2)
                    undamaged = sh.cell(row=i, column=3)
                    working = sh.cell(row=i, column=4)
                    comment = sh.cell(row=i, column=6)

                    #Handle outside of the condition table
                    if sor == title.value:
                        is_start = True
                        continue
                    elif not is_start:
                        if title.value == 'Property':
                            entry.set_property_address(clean.value)
                        elif title.value == 'Date of Inspection':
                            entry.set_date(str(clean.value).strip().split()[0], '%Y-%m-%d')
                    elif eor == title.value:
                        entry.attach_room_and_items()
                        break

                    #iterate through table
                    if is_start:                
                        if title.font.bold and clean.value == 'C':
                            main_title, alt_title = entry.get_alt_title(title.value)
                            if entry.room:
                                entry.attach_room_and_items()
                                entry.create_room(main_title, alt_title, None)
                                
                            else:
                                entry.create_room(main_title, alt_title, None)
                            
                        else:
                            if title.value == None and comment.value and entry.items:
                                entry.combine_comments_in_item(comment)
                            
                            elif comment.value == None and title.value == None or title.value =='C - Clean U - Undamaged W - Working Y - Yes N - No':
                                continue
                            else:
                                entry.add_item(title, clean, undamaged, working,comment)
                        
                        
                filename = os.path.split(file)[1] # Name of the file
                wo_ext = os.path.splitext(filename)[0] # Name without .xlsx
                entry.create_json(str(wo_ext))
                entry.insert_to_db(str(wo_ext) + f' {entry.property_adress}')
                wb.close()