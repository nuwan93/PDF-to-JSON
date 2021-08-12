from re import split
import sys
sys.path.insert(0, r'C:\Users\Nuwan\Downloads\pdf_to_json-20210617T052048Z-001\pdf_to_json\report_conversion')
import Entry 
import openpyxl
from openpyxl.cell import cell
import os
from dateutil.parser import parse


Entry.Entry.load_excel_files(r'E:\001\work\AssetOwl\Atree_word_to_json\docs', 'c5d3cea8-fe7e-4b62-81e8-320ef92fa274')

#start and end of the report
eor = 'Approximate dates when work was last done on residential premises:'


for file in Entry.Entry.files:
    wb = openpyxl.load_workbook(rf'{file}')
    sh = wb.active

    entry = Entry.Entry('ENTRY')
    is_start = False

    # iterate through excel and extract data
    for i in range(3, sh.max_row+1):
        title = sh.cell(row=i, column=1)
        comment = sh.cell(row=i, column=2)
        clean = sh.cell(row=i, column=3)
        undamaged = sh.cell(row=i, column=4)
        working = sh.cell(row=i, column=5)
        
        #Handle outside of the condition table
        if title.fill.start_color.rgb == 'FF000000':
            is_start = True
            continue
        elif not is_start:
            if title.value == '    ADDRESS:  ':
                entry.set_property_address(comment.value)
            elif title.value == '      LEASE START:':
                lease_date = clean.value
                if not lease_date:
                    continue
                lease_date = parse(str(lease_date).strip())
                entry.set_date(str(lease_date).split()[0], '%Y-%m-%d')
        elif eor == title.value:
            entry.attach_room_and_items()
            break

        #iterate through table
        if is_start:                
            if title.fill.start_color.rgb == 'FFBFBFBF' :
                if entry.is_same_room(title.value):
                    continue
                main_title, alt_title = entry.get_alt_title(title.value)
                if entry.room:
                    entry.attach_room_and_items()
                    entry.create_room(main_title, alt_title, None)
                    
                else:
                    entry.create_room(main_title, alt_title, None)
                
            else:
                if title.value == None and comment.value and entry.items:
                    entry.combine_comments_in_item(comment)
                
                elif comment.value == None:
                    continue
                else:
                    entry.add_item(title, clean, undamaged, working,comment)
            
            
    filename = os.path.split(file)[1] # Name of the file
    wo_ext = os.path.splitext(filename)[0] # Name without .xlsx
    entry.create_json(str(wo_ext))
    entry.insert_to_db(str(wo_ext) + f' {entry.property_adress}')
    wb.close()