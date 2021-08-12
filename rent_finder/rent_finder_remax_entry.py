from re import split
import sys
sys.path.insert(0, r'C:\Users\Nuwan\Downloads\pdf_to_json-20210617T052048Z-001\pdf_to_json\report_conversion')
import Entry 
import openpyxl
from openpyxl.cell import cell
import os
from dateutil.parser import parse


Entry.Entry.load_excel_files(r'E:\001\work\AssetOwl\rentfinder PCR - PDF to JSON\word', 'c5d3cea8-fe7e-4b62-81e8-320ef92fa274')

#start and end of the report
sor = 'C – Clean     U – Undamaged     W – Working'



for file in Entry.Entry.files:
    wb = openpyxl.load_workbook(rf'{file}')
    sh = wb.active

    entry = Entry.Entry('ENTRY')
    is_start = False
    address = ''

    # iterate through excel and extract data
    for i in range(2, sh.max_row+1):
        title = sh.cell(row=i, column=1)
        clean = sh.cell(row=i, column=2)
        undamaged = sh.cell(row=i, column=3)
        working = sh.cell(row=i, column=4)
        comment = sh.cell(row=i, column=5)

        #Handle outside of the condition table
        if sor == title.value:
            is_start = True
        elif not is_start:
            if title.font.size == 18:
                address +=  f' {title.value}'
            elif title.value == 'Date of Inspection:':
                inpection_date = parse(str(clean.value).strip())
                entry.set_date(str(inpection_date).split()[0], '%Y-%m-%d')

        #iterate through table
        if is_start:                
            if title.font.size == 11 and clean.value == 'C':
                main_title, alt_title = entry.get_alt_title(title.value)
                if entry.room:
                    entry.attach_room_and_items()
                    entry.create_room(main_title, alt_title, None)
                    
                else:
                    entry.create_room(main_title, alt_title, None)
                
            else:
                if title.value == None and comment.value and entry.items:
                    entry.combine_comments_in_item(comment)
                
                elif comment.value == None or title.font.bold:
                    continue
                else:
                    entry.add_item(title, clean, undamaged, working,comment)

    print(address)
    entry.attach_room_and_items()        
    entry.set_property_address(address)       
    filename = os.path.split(file)[1] # Name of the file
    wo_ext = os.path.splitext(filename)[0] # Name without .xlsx
    entry.create_json(str(wo_ext))
    entry.insert_to_db(str(wo_ext) + f' {entry.property_adress}')
    wb.close()