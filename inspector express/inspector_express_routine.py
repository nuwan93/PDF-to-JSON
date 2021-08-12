from Routine import Routine
import openpyxl
from datetime import datetime
import time
import os

Routine.load_excel_files(r'D:\Work\AssetOwl\PDF to CSV Inspector express\routine', 'c5d3cea8-fe7e-4b62-81e8-320ef92fa274')

for file in Routine.files:
    wb = openpyxl.load_workbook(rf'{file}')
    sh = wb.active

    date_insheet, name = wb.sheetnames[0].split('(')
    d = datetime.strptime(date_insheet.strip(), '%Y-%m-%d').date()
    unix_time = time.mktime(d.timetuple())

    routine = Routine(unix_time,'routine')

    routine.set_property_address(sh.cell(row=7, column=1).value)

    for i in range(1, sh.max_row+1):
        title = sh.cell(row=i, column=1)
        comment = sh.cell(row=i , column=2)
        
        if sh.max_row == i:
            routine.attach_room()

        if title.border.bottom.style == None and title.border.right.style == None:
            continue
        elif title.font.bold:
            main_title, alt_title = routine.get_alt_title(title.value)
            if routine.room:
                routine.attach_room()
                routine.create_room(main_title, alt_title, sh.cell(row=i + 1 , column=2).value)
            else:
                routine.create_room(main_title, alt_title, sh.cell(row=i + 1 , column=2).value)

        elif not title.font.bold and comment.value and routine.room and not title.value == 'Overall':
            routine.combine_comments_in_room(comment)
    
    filename = os.path.split(file)[1] # Name of the file
    wo_ext = os.path.splitext(filename)[0] # Name without .xlsx
    json_file_name = str(wo_ext) + f' {routine.property_adress}'
    routine.create_json(json_file_name)
    routine.insert_to_db(str(wo_ext) + f' {routine.property_adress}')
    wb.close()