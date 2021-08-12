from Entry import Entry
import openpyxl
from openpyxl.cell import cell
from datetime import datetime
import time
import os

Entry.load_excel_files(r'D:\Work\AssetOwl\PDF to CSV Inspector express\Entry', 'c5d3cea8-fe7e-4b62-81e8-320ef92fa274')

for file in Entry.files:
    wb = openpyxl.load_workbook(rf'{file}')
    sh = wb.active

    date_insheet, name = wb.sheetnames[0].split('(')
    d = datetime.strptime(date_insheet.strip(), '%Y-%m-%d').date()
    unix_time = time.mktime(d.timetuple())

    entry = Entry(unix_time, 'ENTRY' )

    # iterate through excel and display data
    for i in range(3, sh.max_row+1):
        title = sh.cell(row=i, column=1)
        clean = sh.cell(row=i, column=2)
        undamaged = sh.cell(row=i, column=3)
        working = sh.cell(row=i, column=4)
        comment = sh.cell(row=i, column=5)

        if sh.max_row == i:
            entry.attach_room_and_items()

        elif (title.border.bottom.style == None and title.border.right.style == None) or title.value == 'Agent section':
            if title.value == 'Property Address:':
                entry.set_property_address(clean.value)
            continue

        elif title.font.bold and title.border:
            main_title, alt_title = entry.get_alt_title(title.value)
            if entry.room:
                entry.attach_room_and_items()
                entry.create_room(main_title, alt_title, None)
                
            else:
                entry.create_room(main_title, alt_title, None)
            
        else:
            if title.value == None and not comment.font.italic  and entry.items:
                entry.combine_comments_in_item(comment)
            
            elif comment.font.italic:
                continue
            else:
                entry.add_item(title, clean, undamaged, working,comment)
        
        
            
    filename = os.path.split(file)[1] # Name of the file
    wo_ext = os.path.splitext(filename)[0] # Name without .xlsx
    json_file_name = str(wo_ext) + f' {entry.property_adress}'
    entry.create_json(json_file_name)
    entry.insert_to_db(str(wo_ext) + f' {entry.property_adress}')
    wb.close()